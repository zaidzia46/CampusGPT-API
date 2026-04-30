import json
import math
import pandas as pd
from datetime import datetime, timezone
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SEMESTER = "spring_2026"
SOURCE   = "fee_structure"

SRC_FILE = "/mnt/user-data/uploads/fee_structure.xlsx"
OUT_XL   = "/home/claude/cui_rag/data/sources/fee_structure_spring_2026.xlsx"
OUT_JSON = "/home/claude/cui_rag/data/chunks/fee_chunks_spring_2026.json"


def fmt_rs(v):
    """Format a numeric rupee value nicely, or return string as-is."""
    if v is None:
        return "N/A"
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, float) and math.isnan(v):
        return "N/A"
    return f"{int(v):,}"

def is_lumpsum(v):
    return isinstance(v, str) and "lumpsum" in v.lower()

def has_admission_fee(row):
    v = row.get("admission_fee")
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return False
    return True

def program_level(name: str) -> str:
    n = name.strip().upper()
    if n.startswith("PHD"):  return "PhD"
    if n.startswith("MS"):   return "MS"
    return "BS"

def program_category(name: str) -> str:
    n = name.lower()
    if any(x in n for x in ("software", "computer")):          return "engineering_cs"
    if any(x in n for x in ("food", "biochem", "biotech",
                             "bioinformatics", "agriculture",
                             "bioscience", "microbiology")):    return "life_sciences"
    if any(x in n for x in ("business", "management")):        return "business"
    if "english" in n:                                          return "humanities"
    if "math" in n:                                             return "mathematics"
    return "general"

def narrative(r: dict) -> str:
    name    = r["program"]
    level   = program_level(name)
    adm_fee = r.get("admission_fee")
    sem_fee = r.get("semester_fee")
    lump    = is_lumpsum(sem_fee)

    if lump:
        lump_clean = str(sem_fee).replace("(Lumpsum)","").replace("(lumpsum)","").strip()
        try:
            lump_clean = f"{int(float(lump_clean)):,}"
        except Exception:
            pass
        fee_line = (
            f"The program operates on a lumpsum fee model. The total fee payable is "
            f"Rs. {lump_clean} as a lumpsum amount. No separate admission fee is "
            f"charged — the lumpsum covers all dues. This lumpsum model applies because "
            f"the program is covered under a special internal scholarship that waives "
            f"standard per-semester charges."
        )
    elif has_admission_fee(r):
        fee_line = (
            f"The semester fee is Rs. {fmt_rs(sem_fee)} per semester. In addition, a "
            f"one-time admission fee of Rs. {fmt_rs(adm_fee)} is charged at the time of "
            f"admission. The admission fee is a non-recurring charge paid only once "
            f"throughout the degree."
        )
    else:
        fee_line = (
            f"The semester fee is Rs. {fmt_rs(sem_fee)} per semester. No separate "
            f"admission fee is applicable for this program."
        )

    return (
        f"{name} is a {level}-level program offered at COMSATS University Islamabad, "
        f"Sahiwal Campus. {fee_line} "
        f"Fee figures are for Spring 2026 and are subject to revision by CUI, "
        f"PS Islamabad. Students should confirm the current fee with the Accounts "
        f"Office or Student Support Center before each semester."
    )

def faq(r: dict) -> str:
    name    = r["program"]
    adm_fee = r.get("admission_fee")
    sem_fee = r.get("semester_fee")
    lump    = is_lumpsum(sem_fee)

    lines = []

    sem_num_clean = None if (sem_fee is None or lump) else int(sem_fee)
    adm_num_clean = None if not has_admission_fee(r) else int(adm_fee)

    if lump:
        lump_amount = fmt_rs(sem_fee).replace(" (Lumpsum)","").replace("(Lumpsum)","").strip()
        lines += [
            f"Q: What is the fee for {name}?",
            f"A: {name} has a lumpsum fee of Rs. {lump_amount}. "
            f"No additional admission fee is charged.",
            "",
            f"Q: Why is {name} on a lumpsum model instead of a per-semester fee?",
            f"A: These programs are covered under a special internal scholarship that "
            f"restructures the fee as a single lumpsum amount rather than a standard "
            f"semester fee plus admission fee.",
            "",
            f"Q: Does {name} have an admission fee?",
            f"A: No. The lumpsum fee covers all charges.",
        ]
    else:
        lines += [
            f"Q: What is the semester fee for {name}?",
            f"A: The semester fee is Rs. {sem_num_clean:,} per semester for Spring 2026.",
            "",
        ]
        if has_admission_fee(r):
            lines += [
                f"Q: Is there an admission fee for {name}?",
                f"A: Yes. A one-time admission fee of Rs. {adm_num_clean:,} is charged at "
                f"the time of admission. It is paid only once and is non-refundable.",
                "",
                f"Q: What is the total cost in the first semester of {name}?",
                f"A: In the first semester you pay Rs. {sem_num_clean:,} (semester fee) "
                f"+ Rs. {adm_num_clean:,} (one-time admission fee) = "
                f"Rs. {sem_num_clean + adm_num_clean:,} total.",
                "",
            ]
        else:
            lines += [
                f"Q: Is there a separate admission fee for {name}?",
                f"A: No. No separate admission fee is charged for this program.",
                "",
            ]

    lines += [
        f"Q: Can the fee for {name} change?",
        f"A: Yes. Fee figures are for Spring 2026 and are subject to revision by CUI, "
        f"PS Islamabad at any stage. Always confirm with the Accounts Office.",
    ]

    return "\n".join(lines)

def meta_text(r: dict) -> str:
    return (
        f"Program: {r['program']} | Level: {program_level(r['program'])} | "
        f"Category: {program_category(r['program'])} | "
        f"Semester fee: Rs. {fmt_rs(r.get('semester_fee'))} | "
        f"Admission fee: Rs. {fmt_rs(r.get('admission_fee'))} | "
        f"Lumpsum model: {is_lumpsum(r.get('semester_fee'))} | "
        f"Source: {SOURCE} | Semester: {SEMESTER}"
    )

def make_metadata(r: dict) -> dict:
    sem_fee = r.get("semester_fee")
    adm_fee = r.get("admission_fee")
    return {
        "program":            r["program"],
        "level":              program_level(r["program"]),
        "category":           program_category(r["program"]),
        "semester_fee_rs":    None if (sem_fee is None or is_lumpsum(sem_fee)) else int(sem_fee),
        "admission_fee_rs":   None if not has_admission_fee(r) else int(adm_fee),
        "lumpsum":            is_lumpsum(sem_fee),
        "lumpsum_amount_rs":  fmt_rs(sem_fee) if is_lumpsum(sem_fee) else None,
        "source":             SOURCE,
        "semester":           SEMESTER,
    }

POLICY_CHUNKS = [
    {
    "chunk_id":   f"{SOURCE}_{SEMESTER}_policy_admission_fee_general",
    "chunk_type": "policy",
    "topic":      SOURCE,
    "semester":   SEMESTER,
    "source":     SOURCE,
    "text": (
        "The admission fee at COMSATS University Islamabad, Sahiwal Campus is "
        "Rs. 22,000. This is a one-time payment charged only at the time of "
        "admission — it is not charged again in subsequent semesters. "
        "Programs under the Special Scholarship scheme (BS Bioinformatics, "
        "BS Business Administration, BS Mathematics, BS Mathematics with Data "
        "Science, and BS English) do not have a separate admission fee — they "
        "use a lumpsum fee model of Rs. 86,000 instead."
    ),
    "metadata": {"source": SOURCE, "semester": SEMESTER,
                 "chunk_type": "policy", "topic": "admission_fee_general"},
},
    {
        "chunk_id":   f"{SOURCE}_{SEMESTER}_policy_admission_fee",
        "chunk_type": "policy",
        "topic":      SOURCE,
        "semester":   SEMESTER,
        "source":     SOURCE,
        "text": (
            "At COMSATS University Islamabad, Sahiwal Campus, a one-time admission fee of "
            "Rs. 22,000 is charged at the time of admission for most programs. This is a "
            "non-recurring charge — it is paid only once and does not apply to subsequent "
            "semesters. Programs under the Special Scholarship scheme (BS Bioinformatics, "
            "BS Business Administration, BS Mathematics, BS Mathematics with Data Science, "
            "and BS English) do not charge a separate admission fee; instead they operate "
            "on a lumpsum fee model of Rs. 86,000."
        ),
        "metadata": {"source": SOURCE, "semester": SEMESTER,
                     "chunk_type": "policy", "topic": "admission_fee_policy"},
    },
    {
        "chunk_id":   f"{SOURCE}_{SEMESTER}_policy_lumpsum",
        "chunk_type": "policy",
        "topic":      SOURCE,
        "semester":   SEMESTER,
        "source":     SOURCE,
        "text": (
            "Certain BS programs at COMSATS University Islamabad, Sahiwal Campus use a "
            "lumpsum fee model instead of a standard semester fee plus admission fee "
            "structure. These programs are: BS Bioinformatics, BS Business Administration, "
            "BS Mathematics, BS Mathematics with Data Science, and BS English. For these "
            "programs, the total fee is Rs. 86,000 as a lumpsum — no additional admission "
            "fee is charged. This model exists because these programs are covered under a "
            "special internal scholarship that restructures the payment."
        ),
        "metadata": {"source": SOURCE, "semester": SEMESTER,
                     "chunk_type": "policy", "topic": "lumpsum_fee_policy"},
    },
    {
        "chunk_id":   f"{SOURCE}_{SEMESTER}_policy_fee_revision",
        "chunk_type": "policy",
        "topic":      SOURCE,
        "semester":   SEMESTER,
        "source":     SOURCE,
        "text": (
            "All fee figures for COMSATS University Islamabad, Sahiwal Campus are specific "
            "to Spring 2026 and are subject to revision by CUI, PS Islamabad at any stage. "
            "Students should confirm the current fee with the Accounts Office or Student "
            "Support Center before each semester. Fee structures may change between semesters."
        ),
        "metadata": {"source": SOURCE, "semester": SEMESTER,
                     "chunk_type": "policy", "topic": "fee_revision_disclaimer"},
    },
    {
        "chunk_id":   f"{SOURCE}_{SEMESTER}_policy_fee_overview",
        "chunk_type": "policy",
        "topic":      SOURCE,
        "semester":   SEMESTER,
        "source":     SOURCE,
        "text": (
            "Fee structure overview for COMSATS University Islamabad, Sahiwal Campus, "
            "Spring 2026. BS Engineering / CS programs (BS Software Engineering, BS Computer "
            "Science): Rs. 133,000 per semester + Rs. 22,000 admission fee. "
            "BS Life Sciences programs (BS Food Science & Nutrition, BS Biochemistry, "
            "BS Biotechnology, BS Agriculture): Rs. 123,000 per semester + Rs. 22,000 "
            "admission fee. Special scholarship BS programs (BS Bioinformatics, BS Business "
            "Administration, BS Mathematics, BS Mathematics with Data Science, BS English): "
            "Rs. 86,000 lumpsum, no admission fee. MS programs (all disciplines): "
            "Rs. 77,000 per semester + Rs. 22,000 admission fee. PhD programs (all "
            "disciplines): Rs. 83,000 per semester + Rs. 22,000 admission fee."
        ),
        "metadata": {"source": SOURCE, "semester": SEMESTER,
                     "chunk_type": "policy", "topic": "fee_overview_all_programs"},
    },
]

def build_chunks(df: pd.DataFrame) -> list[dict]:
    import math

    # Convert fee columns from string to numeric
    for col in ("admission_fee", "semester_fee", "total_first_semester", "sr_no"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    chunks = list(POLICY_CHUNKS)

    for _, row in df.iterrows():
        r = {k: (None if (isinstance(v, float) and math.isnan(v)) else v)
             for k, v in row.items()}

        program = r.get("program", "unknown")
        if not program:
            continue

        # Build a clean slug for the chunk ID
        slug = str(program).lower()\
                           .replace(" ", "_")\
                           .replace("&", "and")\
                           .replace("(", "")\
                           .replace(")", "")\
                           .replace("/", "_")
        base_id = f"{SOURCE}_{SEMESTER}_{slug}"

        try:
            chunks += [
                {
                    "chunk_id":   f"{base_id}_narrative",
                    "chunk_type": "narrative",
                    "topic":      SOURCE,
                    "semester":   SEMESTER,
                    "source":     SOURCE,
                    "text":       narrative(r),
                    "metadata":   make_metadata(r),
                },
                {
                    "chunk_id":   f"{base_id}_faq",
                    "chunk_type": "faq",
                    "topic":      SOURCE,
                    "semester":   SEMESTER,
                    "source":     SOURCE,
                    "text":       faq(r),
                    "metadata":   make_metadata(r),
                },
                {
                    "chunk_id":   f"{base_id}_metadata",
                    "chunk_type": "metadata",
                    "topic":      SOURCE,
                    "semester":   SEMESTER,
                    "source":     SOURCE,
                    "text":       meta_text(r),
                    "metadata":   make_metadata(r),
                },
            ]
        except Exception as e:
            print(f"  [WARN] Skipped {program}: {e}")

    return chunks

def build_admin_excel(df: pd.DataFrame, out_path: str):
    wb = Workbook()

    C_HDR_BG  = "1F4E79"
    C_HDR_FG  = "FFFFFF"
    C_BS_ENG  = "DEEAF1"   # blue  — BS Eng/CS
    C_BS_SCI  = "E2EFDA"   # green — BS Life Sciences
    C_BS_SPEC = "FFF2CC"   # amber — BS Special/Lumpsum
    C_MS      = "F4CCFF"   # purple — MS
    C_PHD     = "FFE0CC"   # coral  — PhD
    C_SEC_BG  = "2E75B6"
    C_SEC_FG  = "FFFFFF"
    C_ALT     = "F5F9FF"
    C_BRD     = "BFBFBF"

    thin   = Side(style="thin",   color=C_BRD)
    medium = Side(style="medium", color="1F4E79")
    bt = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hf(bold=True, size=10, color=C_HDR_FG):
        return Font(name="Arial", bold=bold, size=size, color=color)
    def cf(bold=False, size=10, color="000000"):
        return Font(name="Arial", bold=bold, size=size, color=color)
    def fl(hex_c):
        return PatternFill("solid", fgColor=hex_c)
    def ctr():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def lft():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)
    def rgt():
        return Alignment(horizontal="right", vertical="center")

    ws = wb.active
    ws.title = "Fee Structure"

    # Title
    ws.merge_cells("A1:I1")
    ws["A1"] = "COMSATS University Islamabad, Sahiwal Campus — Fee Structure Admin Dashboard"
    ws["A1"].font      = Font(name="Arial", bold=True, size=13, color=C_HDR_FG)
    ws["A1"].fill      = fl(C_HDR_BG)
    ws["A1"].alignment = ctr()
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:I2")
    ws["A2"] = "Source of Truth — Spring 2026   |   Edit this file, then run: python generate_chunks.py --semester spring_2026"
    ws["A2"].font      = Font(name="Arial", italic=True, size=9, color="595959")
    ws["A2"].fill      = fl("D9E2F3")
    ws["A2"].alignment = ctr()
    ws.row_dimensions[2].height = 16

    headers = [
        ("A", "Sr No",                       7),
        ("B", "Program Name",               36),
        ("C", "Level",                        8),
        ("D", "Category",                    18),
        ("E", "Admission Fee\n(Rs.)",        16),
        ("F", "Semester Fee\n(Rs.)",         16),
        ("G", "Lumpsum\nModel",              12),
        ("H", "Total 1st Semester\n(Rs.)",   18),
        ("I", "Notes",                       38),
    ]
    for col_letter, label, width in headers:
        c = ws[f"{col_letter}3"]
        c.value     = label
        c.font      = hf(size=9)
        c.fill      = fl(C_HDR_BG)
        c.alignment = ctr()
        c.border    = bt
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[3].height = 30

    def row_color(name):
        n = name.lower()
        if program_level(name) == "MS":  return C_MS
        if program_level(name) == "PhD": return C_PHD
        if is_lumpsum(None):             return C_BS_SPEC
        if any(x in n for x in ("software","computer")): return C_BS_ENG
        if any(x in n for x in ("food","biochem","biotech","bio","agriculture","micro")): return C_BS_SCI
        return C_BS_SPEC

    sections = {
        "BS Software Engineering": ("__SECTION__", "BS Engineering & Computer Science Programs"),
        "BS Food Science & Nutrition": ("__SECTION__", "BS Life Sciences Programs"),
        "BS Bioinformatics": ("__SECTION__", "BS Special Scholarship Programs (Lumpsum Fee)"),
        "MS Management Sciences": ("__SECTION__", "MS Programs"),
        "PHD Management Sciences": ("__SECTION__", "PhD Programs"),
    }

    excel_row = 4
    alt = False
    for sr, (_, row) in enumerate(df.iterrows(), start=1):
        name = row["program"]

        if name in sections:
            _, sec_label = sections[name]
            ws.merge_cells(f"A{excel_row}:I{excel_row}")
            c = ws[f"A{excel_row}"]
            c.value     = sec_label
            c.font      = Font(name="Arial", bold=True, size=9, color=C_SEC_FG)
            c.fill      = fl(C_SEC_BG)
            c.alignment = lft()
            c.border    = bt
            ws.row_dimensions[excel_row].height = 16
            excel_row += 1
            alt = False

        adm_fee  = row["admission_fee"]
        sem_fee  = row["semester_fee"]
        lump     = is_lumpsum(sem_fee)
        adm_num  = None if (adm_fee is None or (isinstance(adm_fee, float) and math.isnan(adm_fee))) else int(adm_fee)
        sem_num  = None if lump else (None if sem_fee is None else int(sem_fee))
        total_1st = None if lump else (
            (sem_num + adm_num) if (sem_num and adm_num) else sem_num
        )

        row_bg = C_ALT if alt else "FFFFFF"
        # Override with category colour
        n = name.lower()
        if program_level(name) == "MS":          row_bg = "F9F0FF"
        elif program_level(name) == "PhD":        row_bg = "FFF3EC"
        elif lump:                                row_bg = "FFFBE6"
        elif any(x in n for x in ("software","computer")): row_bg = C_ALT if alt else "EEF5FB"
        elif any(x in n for x in ("food","biochem","biotech","bioinformatics",
                                   "agriculture","bioscience","micro")): row_bg = C_ALT if alt else "EFF7EC"
        alt = not alt

        values = {
            "A": sr,
            "B": name,
            "C": program_level(name),
            "D": program_category(name).replace("_"," ").title(),
            "E": adm_num,
            "F": sem_num,
            "G": "Yes" if lump else "No",
            "H": fmt_rs(sem_fee) if lump else (total_1st if total_1st else ""),
            "I": f"Lumpsum: Rs. {fmt_rs(sem_fee)}" if lump else "",
        }

        for col_letter, val in values.items():
            c = ws[f"{col_letter}{excel_row}"]
            c.value  = val
            c.font   = cf(size=9)
            c.fill   = fl(row_bg)
            c.border = bt
            c.alignment = rgt() if col_letter in ("A","E","F","H") else (
                ctr() if col_letter in ("C","G") else lft()
            )

        ws.row_dimensions[excel_row].height = 20
        excel_row += 1

    ws.freeze_panes = "B4"
    ws.auto_filter.ref = f"A3:I{excel_row-1}"

    wl = wb.create_sheet("Legend & Notes")
    wl.merge_cells("A1:C1")
    wl["A1"] = "Legend & Fee Notes — Spring 2026"
    wl["A1"].font      = Font(name="Arial", bold=True, size=11, color=C_HDR_FG)
    wl["A1"].fill      = fl(C_HDR_BG)
    wl["A1"].alignment = ctr()
    wl.row_dimensions[1].height = 22
    for col, w in [("A",20),("B",30),("C",65)]:
        wl.column_dimensions[col].width = w

    legend_data = [
        ("Colour", "Program Group", "Notes"),
        ("Blue row",    "BS Engineering / CS",       "Rs. 133,000/sem + Rs. 22,000 admission fee"),
        ("Green row",   "BS Life Sciences",           "Rs. 123,000/sem + Rs. 22,000 admission fee"),
        ("Amber row",   "BS Special / Lumpsum",       "Rs. 86,000 lumpsum. No admission fee. Special scholarship programs."),
        ("Purple row",  "MS Programs",                "Rs. 77,000/sem + Rs. 22,000 admission fee"),
        ("Coral row",   "PhD Programs",               "Rs. 83,000/sem + Rs. 22,000 admission fee"),
        ("—",           "Admission Fee",              "One-time charge at admission. Not repeated in subsequent semesters."),
        ("—",           "Lumpsum Model",              "Covers all dues. No separate admission fee charged."),
        ("—",           "Fee Validity",               "Spring 2026 only. Subject to revision at CUI, PS Islamabad."),
    ]
    for i, (a, b, c_val) in enumerate(legend_data, start=2):
        is_hdr = i == 2
        bg = C_HDR_BG if is_hdr else ("FFFFFF" if i % 2 == 0 else C_ALT)
        fg = C_HDR_FG if is_hdr else "000000"
        for col, val in zip("ABC", [a, b, c_val]):
            cell = wl[f"{col}{i}"]
            cell.value     = val
            cell.font      = Font(name="Arial", bold=is_hdr, size=9, color=fg)
            cell.fill      = fl(bg)
            cell.alignment = lft()
            cell.border    = bt
        wl.row_dimensions[i].height = 22

    # ── Sheet 3: Semester Log ──
    wlog = wb.create_sheet("Semester Log")
    wlog.merge_cells("A1:E1")
    wlog["A1"] = "Semester Update Log"
    wlog["A1"].font      = Font(name="Arial", bold=True, size=11, color=C_HDR_FG)
    wlog["A1"].fill      = fl(C_HDR_BG)
    wlog["A1"].alignment = ctr()
    wlog.row_dimensions[1].height = 22
    for col, label, width in [("A","Semester",16),("B","Updated By",20),
                               ("C","Date",14),("D","Changes Made",35),("E","Notes",50)]:
        c = wlog[f"{col}2"]
        c.value     = label
        c.font      = hf(size=9)
        c.fill      = fl(C_HDR_BG)
        c.alignment = ctr()
        c.border    = bt
        wlog.column_dimensions[col].width = width
    wlog.row_dimensions[2].height = 20
    for col, val in zip("ABCDE", ["Spring 2026","Admin","2026-01-01",
                                   "fee_structure_spring_2026.xlsx","Initial entry from source file."]):
        c = wlog[f"{col}3"]
        c.value     = val
        c.font      = cf(size=9)
        c.fill      = fl("FFFFFF")
        c.alignment = lft()
        c.border    = bt
    wlog.row_dimensions[3].height = 20

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"  Admin Excel saved → {out_path}")

def build(excel_path: str = None) -> list[dict]:
    """
    Reads fee structure data from Google Sheets and returns all chunks.
    excel_path is kept as parameter for compatibility but is ignored.
    """
    from sheets_reader import read_sheet

    df = read_sheet(
        topic      = "fees",
        sheet_name = "Fee Structure",
        header_row = 2
    )

    # Clean column names
    df.columns = [c.strip().lower().replace(" ", "_").replace("\n", "_")
                  for c in df.columns]
    
    # Drop section separator rows
    first_col = df.columns[0]
    df = df[pd.to_numeric(df[first_col], errors="coerce").notna()].copy()
    df = df.reset_index(drop=True)


    rename = {}
    for c in df.columns:
        if c == "program_name":                        rename[c] = "program"
        elif "admission_fee" in c:                     rename[c] = "admission_fee"
        elif c.startswith("semester_fee"):             rename[c] = "semester_fee"
        elif "total_1st" in c or "total_first" in c:  rename[c] = "total_first_semester"
    df = df.rename(columns=rename)


    df = df[df["program"].notna()].reset_index(drop=True)
    print("[DEBUG] Columns going into build_chunks:", list(df.columns))
    print("[DEBUG] Sample row:")
    for col in df.columns:
        print(f"  {col}: {df.iloc[0].get(col) if len(df) > 0 else 'NO ROWS'}")
    return build_chunks(df)

if __name__ == "__main__":
    # Read source
    df = pd.read_excel(SRC_FILE, dtype={"Semester Fee": str})
    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]
    df = df.rename(columns={"semester_fee": "semester_fee", "admission_fee": "admission_fee"})
    df = df[df["program"].notna()]

    print(f"  Loaded {len(df)} programs from {SRC_FILE}")
    print("[DEBUG] Fee sheet columns:", list(df.columns))

    # Build admin Excel
    build_admin_excel(df, OUT_XL)

    # Build chunks
    chunks = build_chunks(df)

    output = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "semester":     SEMESTER,
        "total_chunks": len(chunks),
        "chunks":       chunks,
    }
    Path(OUT_JSON).parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"  Chunks saved  -> {OUT_JSON}")
    print(f"\n  Summary:")
    types = {}
    for c in chunks:
        t = c["chunk_type"]
        types[t] = types.get(t, 0) + 1
    for t, n in sorted(types.items()):
        print(f"    {t:12}: {n}")
    print(f"    {'TOTAL':12}: {len(chunks)}")
