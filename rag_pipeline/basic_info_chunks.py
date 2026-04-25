"""
run_basic_info_chunks.py
------------------------
Reads Basic_Info.docx content and produces:
  - data/sources/basic_info_spring_2026.xlsx   (admin dashboard)
  - data/chunks/basic_info_spring_2026.json    (RAG-ready chunks)

Unlike scholarships/fees which are tabular, basic info is topic-based.
Each topic gets its own narrative + FAQ + metadata chunk set.
"""

import json
from datetime import datetime, timezone
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SEMESTER = "spring_2026"
SOURCE   = "basic_info"
OUT_XL   = "/home/claude/cui_rag/data/sources/basic_info_spring_2026.xlsx"
OUT_JSON = "/home/claude/cui_rag/data/chunks/basic_info_spring_2026.json"

# ── Raw topic data ─────────────────────────────────────────────────────────
# Each entry: (topic_id, topic_label, category, content_dict)

TOPICS = [

    {
        "topic_id":    "introduction",
        "topic_label": "Campus Introduction & Location",
        "category":    "general_info",
        "facts": {
            "location": "COMSATS Road off G.T Road, Sahiwal — halfway between Lahore and Multan",
            "established": "September 23, 2006",
            "area": "36 acres",
            "admission_url": "https://admissions.comsats.edu.pk/home/procedure",
        },
        "narrative": (
            "COMSATS University Islamabad, Sahiwal Campus is located on COMSATS Road, "
            "off G.T Road, Sahiwal — situated halfway between Lahore and Multan. The campus "
            "was formally inaugurated on September 23, 2006, and is a purposely built facility "
            "spread over 36 acres of land. It is a public sector university operating under the "
            "Ministry of Science and Technology, Government of Pakistan. The official admission "
            "procedure is available online at https://admissions.comsats.edu.pk/home/procedure"
        ),
        "faqs": [
            ("Where is COMSATS University Islamabad, Sahiwal Campus located?",
             "It is located on COMSATS Road, off G.T Road, Sahiwal — halfway between Lahore and Multan."),
            ("When was COMSATS Sahiwal Campus established?",
             "The campus was formally inaugurated on September 23, 2006."),
            ("How large is the COMSATS Sahiwal Campus?",
             "The campus is spread over 36 acres of purposely built land."),
            ("Where can I find the admission procedure for COMSATS?",
             "The admission procedure is available online at https://admissions.comsats.edu.pk/home/procedure"),
        ],
    },

    {
        "topic_id":    "facts_and_figures",
        "topic_label": "Campus Facts & Figures",
        "category":    "general_info",
        "facts": {
            "academic_blocks": 5,
            "faculty_blocks": 1,
            "hostels_outside_campus": 1,
            "total_faculty": 214,
            "phd_faculty": 66,
            "academic_departments": 8,
            "undergraduate_programs": 14,
            "graduate_programs": 9,
            "total_students_enrolled": 3675,
        },
        "narrative": (
            "COMSATS University Islamabad, Sahiwal Campus has 5 academic blocks and 1 faculty "
            "block. There is 1 hostel located outside the campus. The campus has a total of "
            "214 faculty members, of whom 66 hold PhD degrees. There are 8 academic departments "
            "offering 14 undergraduate programs and 9 graduate programs. Total student enrollment "
            "stands at 3,675."
        ),
        "faqs": [
            ("How many students are enrolled at COMSATS Sahiwal Campus?",
             "Total enrollment is 3,675 students."),
            ("How many faculty members does COMSATS Sahiwal have?",
             "There are 214 total faculty members, of whom 66 hold PhD degrees."),
            ("How many programs are offered at COMSATS Sahiwal?",
             "14 undergraduate programs and 9 graduate programs are offered."),
            ("How many academic departments does COMSATS Sahiwal have?",
             "There are 8 academic departments."),
            ("How many blocks does COMSATS Sahiwal Campus have?",
             "There are 5 academic blocks, 1 faculty block, and 1 workshop block (Block W)."),
        ],
    },

    {
        "topic_id":    "undergraduate_programs",
        "topic_label": "Undergraduate Programs Offered",
        "category":    "academics",
        "facts": {
            "programs": [
                "BS Mechanical Engineering",
                "BS Electrical Engineering",
                "BS Civil Engineering",
                "BS Computer Science",
                "BS Telecommunication & Network",
                "BS Software Engineering",
                "BS Bioinformatics",
                "BS Biosciences",
                "BS Food Science & Nutrition",
                "BS Business Administration",
                "BS Accounts & Finance",
            ],
        },
        "narrative": (
            "COMSATS University Islamabad, Sahiwal Campus offers the following undergraduate "
            "(BS-level) programs: Mechanical Engineering, Electrical Engineering, Civil Engineering, "
            "Computer Science, Telecommunication & Network, Software Engineering, Bioinformatics, "
            "Biosciences, Food Science & Nutrition, Business Administration, and Accounts & Finance. "
            "A total of 14 undergraduate programs are offered across 8 academic departments."
        ),
        "faqs": [
            ("What undergraduate programs are offered at COMSATS Sahiwal?",
             "BS programs offered include: Mechanical Engineering, Electrical Engineering, Civil Engineering, "
             "Computer Science, Telecommunication & Network, Software Engineering, Bioinformatics, "
             "Biosciences, Food Science & Nutrition, Business Administration, and Accounts & Finance."),
            ("Does COMSATS Sahiwal offer BS Computer Science?",
             "Yes. BS Computer Science is offered at COMSATS Sahiwal Campus."),
            ("Does COMSATS Sahiwal offer engineering programs?",
             "Yes. BS Mechanical Engineering, BS Electrical Engineering, BS Civil Engineering, "
             "BS Software Engineering, and BS Telecommunication & Network are all offered."),
            ("Is there a business program at COMSATS Sahiwal?",
             "Yes. BS Business Administration and BS Accounts & Finance are both available."),
        ],
    },

    {
        "topic_id":    "graduate_programs",
        "topic_label": "Graduate Programs Offered",
        "category":    "academics",
        "facts": {
            "programs": [
                "MS Mechanical Engineering",
                "MS Management Sciences",
                "MS Computer Science",
                "MS Biosciences",
                "MBA (1.5 years)",
                "MS Mathematics",
            ],
        },
        "narrative": (
            "COMSATS University Islamabad, Sahiwal Campus offers the following graduate programs: "
            "MS Mechanical Engineering, MS Management Sciences, MS Computer Science, MS Biosciences, "
            "MBA (1.5 years), and MS Mathematics. A total of 9 graduate programs are offered. "
            "PhD programs are also available in select disciplines."
        ),
        "faqs": [
            ("What graduate programs are offered at COMSATS Sahiwal?",
             "Graduate programs include: MS Mechanical Engineering, MS Management Sciences, "
             "MS Computer Science, MS Biosciences, MBA (1.5 years), and MS Mathematics."),
            ("Is MBA available at COMSATS Sahiwal?",
             "Yes. An MBA (1.5 years) program is offered at COMSATS Sahiwal Campus."),
            ("Does COMSATS Sahiwal offer MS Computer Science?",
             "Yes. MS Computer Science is available at COMSATS Sahiwal Campus."),
            ("Are PhD programs available at COMSATS Sahiwal?",
             "Yes. PhD programs are offered in select disciplines including Computer Science, "
             "Management Sciences, Mathematics, and Microbiology & Immunology."),
        ],
    },

    {
        "topic_id":    "campus_blocks",
        "topic_label": "Campus Blocks & Department Locations",
        "category":    "campus_facilities",
        "facts": {
            "Block A": "Engineering Department",
            "Block B": "Management Sciences",
            "Block C": "Computer Science",
            "Block D": "Biosciences and Mathematics",
            "Block W": "Workshops",
        },
        "narrative": (
            "COMSATS University Islamabad, Sahiwal Campus is organized into five blocks. "
            "Block A houses the Engineering Department. Block B is for Management Sciences. "
            "Block C is for Computer Science. Block D houses Biosciences and Mathematics. "
            "Block W contains the workshops. There is also a separate faculty block on campus."
        ),
        "faqs": [
            ("Where is the Computer Science department at COMSATS Sahiwal?",
             "The Computer Science department is in Block C."),
            ("Where is the Engineering department at COMSATS Sahiwal?",
             "The Engineering Department is in Block A."),
            ("Where is the Management Sciences department?",
             "Management Sciences is in Block B."),
            ("Where is the Biosciences department at COMSATS Sahiwal?",
             "Biosciences and Mathematics are both located in Block D."),
            ("What is Block W at COMSATS Sahiwal?",
             "Block W contains the workshops."),
        ],
    },

    {
        "topic_id":    "cafeteria_and_printing",
        "topic_label": "Cafeteria & Printing Shop",
        "category":    "campus_facilities",
        "facts": {
            "cafeteria_location": "Center of the campus",
            "cafeteria_items": "Fast food, shakes, traditional dishes",
            "printing_shop_location": "Next to the cafeteria",
        },
        "narrative": (
            "COMSATS University Islamabad, Sahiwal Campus has a cafeteria located in the center "
            "of the campus. It offers a variety of food items including fast food, shakes, and "
            "traditional dishes. A printing shop is located next to the cafeteria. Students can "
            "use the printing shop to print documents, assignments, and other academic materials."
        ),
        "faqs": [
            ("Where is the cafeteria at COMSATS Sahiwal?",
             "The cafeteria is located in the center of the campus."),
            ("What food is available at the COMSATS Sahiwal cafeteria?",
             "The cafeteria offers fast food, shakes, and traditional dishes."),
            ("Where can I print documents on campus?",
             "The printing shop is located next to the cafeteria on campus."),
            ("Is there a canteen or food area at COMSATS Sahiwal?",
             "Yes. There is a cafeteria in the center of the campus offering fast food, shakes, and traditional dishes."),
        ],
    },

    {
        "topic_id":    "parking",
        "topic_label": "Parking Facilities",
        "category":    "campus_facilities",
        "facts": {
            "student_car_parking": "Not available",
            "student_bike_parking": "Available outside campus at Rs. 20",
            "faculty_parking": "Cars and bikes allowed inside university",
        },
        "narrative": (
            "At COMSATS University Islamabad, Sahiwal Campus, students do not have dedicated "
            "car parking. Bike parking is available outside the campus at a fee of Rs. 20. "
            "Faculty members are permitted to park both cars and bikes inside the university premises."
        ),
        "faqs": [
            ("Is there car parking for students at COMSATS Sahiwal?",
             "No. Students do not have dedicated car parking at the campus."),
            ("Is there bike parking for students at COMSATS Sahiwal?",
             "Yes. Bike parking is available outside the campus at Rs. 20."),
            ("Can faculty park inside COMSATS Sahiwal Campus?",
             "Yes. Faculty members can park both cars and bikes inside the university."),
            ("How much does bike parking cost at COMSATS Sahiwal?",
             "Bike parking outside the campus costs Rs. 20."),
        ],
    },

    {
        "topic_id":    "transport",
        "topic_label": "Electric Bus & Transport",
        "category":    "campus_facilities",
        "facts": {
            "service": "Electric bus",
            "frequency": "Every 1 hour and 15 minutes",
            "routes": ["Gamber (before Okara)", "Sahiwal City"],
        },
        "narrative": (
            "An electric bus service operates outside COMSATS University Islamabad, Sahiwal Campus. "
            "The bus runs every 1 hour and 15 minutes and provides routes to Gamber (located before "
            "Okara) and Sahiwal City. This service facilitates commuting for students and staff "
            "traveling to and from the campus."
        ),
        "faqs": [
            ("Is there a bus service at COMSATS Sahiwal Campus?",
             "Yes. An electric bus operates outside the campus every 1 hour and 15 minutes."),
            ("What routes does the COMSATS Sahiwal electric bus cover?",
             "The bus covers routes to Gamber (before Okara) and Sahiwal City."),
            ("How often does the electric bus run at COMSATS Sahiwal?",
             "The bus runs every 1 hour and 15 minutes."),
            ("How can I travel to COMSATS Sahiwal from Sahiwal City?",
             "An electric bus runs from Sahiwal City to the campus every 1 hour and 15 minutes."),
        ],
    },

    {
        "topic_id":    "grounds_and_pandals",
        "topic_label": "Grounds, Sports & Pandals",
        "category":    "campus_facilities",
        "facts": {
            "sports_ground": "Front right corner of campus",
            "small_pandal":  "In front of the Admission Office",
            "main_pandal":   "Backside of the CS Block",
        },
        "narrative": (
            "COMSATS University Islamabad, Sahiwal Campus has several outdoor spaces. "
            "A sports ground is located at the front right corner of the campus. "
            "A small pandal (covered outdoor area) is situated in front of the Admission Office. "
            "The main pandal is located at the backside of the CS Block (Block C)."
        ),
        "faqs": [
            ("Where is the sports ground at COMSATS Sahiwal?",
             "The sports ground is at the front right corner of the campus."),
            ("Where is the main pandal at COMSATS Sahiwal?",
             "The main pandal is at the backside of the CS Block (Block C)."),
            ("Is there an outdoor event space at COMSATS Sahiwal?",
             "Yes. There is a small pandal in front of the Admission Office and a main pandal at the back of the CS Block."),
        ],
    },

    {
        "topic_id":    "hostel",
        "topic_label": "Hostel",
        "category":    "campus_facilities",
        "facts": {
            "count": 1,
            "location": "Outside the campus",
        },
        "narrative": (
            "COMSATS University Islamabad, Sahiwal Campus has one hostel facility, which is "
            "located outside the main campus. Students requiring on-campus accommodation should "
            "contact the Student Support Center or the administration for availability, "
            "allocation procedures, and fee details."
        ),
        "faqs": [
            ("Does COMSATS Sahiwal have a hostel?",
             "Yes. There is one hostel, located outside the main campus."),
            ("Is the hostel inside or outside COMSATS Sahiwal Campus?",
             "The hostel is located outside the campus."),
            ("How many hostels does COMSATS Sahiwal have?",
             "There is one hostel facility located outside the main campus."),
        ],
    },

    {
        "topic_id":    "additional_fees",
        "topic_label": "Additional Fees — Registration, Credit Hour & Degree",
        "category":    "fees",
        "facts": {
            "admission_fee":               "Rs. 22,000 — one-time at admission, in addition to semester fee",
            "extra_registration_fee":      "Applicable if student registers for additional semester",
            "credit_hour_fee_ug_masters":  "Rs. 6,000 per credit hour",
            "credit_hour_fee_ms_phd":      "Rs. 4,000 per credit hour",
            "degree_fee":                  "Rs. 10,000 — charged upon program completion",
        },
        "narrative": (
            "At COMSATS University Islamabad, Sahiwal Campus, the admission fee of Rs. 22,000 "
            "is a one-time payment charged to new students at the time of admission, in addition "
            "to the semester fee. If a student registers for an additional semester beyond the "
            "normal program duration, an extra registration fee and a credit hour fee will apply. "
            "The credit hour fee is Rs. 6,000 per credit hour for undergraduate and Master's "
            "programs, and Rs. 4,000 per credit hour for MS and PhD programs. Upon successful "
            "completion of the degree, a degree fee of Rs. 10,000 is charged."
        ),
        "faqs": [
            ("What is the admission fee at COMSATS Sahiwal?",
             "Rs. 22,000 — a one-time payment charged at admission, in addition to the semester fee."),
            ("What is the credit hour fee at COMSATS Sahiwal?",
             "Rs. 6,000 per credit hour for undergraduate and Master's programs. "
             "Rs. 4,000 per credit hour for MS and PhD programs."),
            ("Is there a degree fee at COMSATS Sahiwal?",
             "Yes. A degree fee of Rs. 10,000 is charged upon completion of the program."),
            ("What extra charges apply if I take an additional semester?",
             "An extra registration fee and a credit hour fee will apply. "
             "Credit hour fee is Rs. 6,000 (UG/Masters) or Rs. 4,000 (MS/PhD) per credit hour."),
            ("Is the Rs. 22,000 admission fee charged every semester?",
             "No. It is a one-time charge paid only at the time of admission."),
        ],
    },
]


# ── Chunk builder ──────────────────────────────────────────────────────────

def build_chunks(topics: list) -> list[dict]:
    chunks = []
    for t in topics:
        base_id = f"{SOURCE}_{SEMESTER}_{t['topic_id']}"

        # Narrative chunk
        chunks.append({
            "chunk_id":   f"{base_id}_narrative",
            "chunk_type": "narrative",
            "topic":      t["topic_id"],
            "semester":   SEMESTER,
            "source":     SOURCE,
            "text":       t["narrative"],
            "metadata": {
                "topic_id":    t["topic_id"],
                "topic_label": t["topic_label"],
                "category":    t["category"],
                "source":      SOURCE,
                "semester":    SEMESTER,
            },
        })

        # FAQ chunk — all Q&A pairs for this topic in one chunk
        faq_text = "\n\n".join(
            f"Q: {q}\nA: {a}" for q, a in t["faqs"]
        )
        chunks.append({
            "chunk_id":   f"{base_id}_faq",
            "chunk_type": "faq",
            "topic":      t["topic_id"],
            "semester":   SEMESTER,
            "source":     SOURCE,
            "text":       faq_text,
            "metadata": {
                "topic_id":    t["topic_id"],
                "topic_label": t["topic_label"],
                "category":    t["category"],
                "source":      SOURCE,
                "semester":    SEMESTER,
            },
        })

        # Metadata chunk — compact tag line for filtered retrieval
        facts_str = " | ".join(
            f"{k}: {v}" for k, v in t["facts"].items()
            if not isinstance(v, list)
        )
        if not facts_str:
            facts_str = "See narrative for details"
        chunks.append({
            "chunk_id":   f"{base_id}_metadata",
            "chunk_type": "metadata",
            "topic":      t["topic_id"],
            "semester":   SEMESTER,
            "source":     SOURCE,
            "text": (
                f"Topic: {t['topic_label']} | Category: {t['category']} | "
                f"{facts_str} | Source: {SOURCE} | Semester: {SEMESTER}"
            ),
            "metadata": {
                "topic_id":    t["topic_id"],
                "topic_label": t["topic_label"],
                "category":    t["category"],
                "source":      SOURCE,
                "semester":    SEMESTER,
            },
        })

    return chunks


# ── Admin Excel builder ────────────────────────────────────────────────────

def build_admin_excel(topics: list, out_path: str):
    wb = Workbook()

    C_HDR_BG    = "1F4E79"
    C_HDR_FG    = "FFFFFF"
    C_GEN       = "DEEAF1"   # blue    — general info
    C_ACAD      = "E2EFDA"   # green   — academics
    C_FAC       = "FFF2CC"   # amber   — campus facilities
    C_FEE       = "F4CCFF"   # purple  — fees
    C_SEC_BG    = "2E75B6"
    C_SEC_FG    = "FFFFFF"
    C_BRD       = "BFBFBF"

    thin = Side(style="thin", color=C_BRD)
    bt   = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hf(bold=True, size=10, color=C_HDR_FG):
        return Font(name="Arial", bold=bold, size=size, color=color)
    def cf(bold=False, size=10, color="000000"):
        return Font(name="Arial", bold=bold, size=size, color=color)
    def fl(hex_c):
        return PatternFill("solid", fgColor=hex_c)
    def ctr():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def lft():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    CAT_COLOR = {
        "general_info":      C_GEN,
        "academics":         C_ACAD,
        "campus_facilities": C_FAC,
        "fees":              C_FEE,
    }

    # ── Sheet 1: Topics ──
    ws = wb.active
    ws.title = "Basic Info Topics"

    ws.merge_cells("A1:G1")
    ws["A1"] = "COMSATS University Islamabad, Sahiwal Campus — Basic Info Admin Dashboard"
    ws["A1"].font      = Font(name="Arial", bold=True, size=13, color=C_HDR_FG)
    ws["A1"].fill      = fl(C_HDR_BG)
    ws["A1"].alignment = ctr()
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    ws["A2"] = "Source of Truth — Spring 2026   |   Edit narrative/FAQ fields, then run: python generate_chunks.py --semester spring_2026"
    ws["A2"].font      = Font(name="Arial", italic=True, size=9, color="595959")
    ws["A2"].fill      = fl("D9E2F3")
    ws["A2"].alignment = ctr()
    ws.row_dimensions[2].height = 16

    headers = [
        ("A", "Sr No",          7),
        ("B", "Topic ID",       26),
        ("C", "Topic Label",    34),
        ("D", "Category",       20),
        ("E", "Narrative Text", 70),
        ("F", "No. of FAQs",    12),
        ("G", "Key Facts Summary", 55),
    ]
    for col_letter, label, width in headers:
        c = ws[f"{col_letter}3"]
        c.value     = label
        c.font      = hf(size=9)
        c.fill      = fl(C_HDR_BG)
        c.alignment = ctr()
        c.border    = bt
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[3].height = 22

    prev_cat = None
    excel_row = 4
    for sr, t in enumerate(topics, start=1):
        cat = t["category"]

        # Section separator when category changes
        if cat != prev_cat:
            sec_label = cat.replace("_", " ").title()
            ws.merge_cells(f"A{excel_row}:G{excel_row}")
            c = ws[f"A{excel_row}"]
            c.value     = f"— {sec_label} —"
            c.font      = Font(name="Arial", bold=True, size=9, color=C_SEC_FG)
            c.fill      = fl(C_SEC_BG)
            c.alignment = lft()
            c.border    = bt
            ws.row_dimensions[excel_row].height = 16
            excel_row += 1
            prev_cat = cat

        row_bg = CAT_COLOR.get(cat, "FFFFFF")

        facts_summary = "; ".join(
            f"{k}: {v}" for k, v in t["facts"].items()
            if not isinstance(v, list)
        )
        if not facts_summary:
            facts_summary = "; ".join(
                str(item) for sublist in t["facts"].values()
                if isinstance(sublist, list) for item in sublist[:3]
            ) + " ..."

        values = {
            "A": sr,
            "B": t["topic_id"],
            "C": t["topic_label"],
            "D": cat.replace("_", " ").title(),
            "E": t["narrative"],
            "F": len(t["faqs"]),
            "G": facts_summary,
        }

        for col_letter, val in values.items():
            c = ws[f"{col_letter}{excel_row}"]
            c.value     = val
            c.font      = cf(size=9)
            c.fill      = fl(row_bg)
            c.border    = bt
            c.alignment = ctr() if col_letter in ("A","F") else lft()

        ws.row_dimensions[excel_row].height = 60
        excel_row += 1

    ws.freeze_panes = "B4"

    # ── Sheet 2: FAQ Details ──
    wf = wb.create_sheet("FAQ Details")
    wf.merge_cells("A1:D1")
    wf["A1"] = "FAQ Details — All Topics"
    wf["A1"].font      = Font(name="Arial", bold=True, size=11, color=C_HDR_FG)
    wf["A1"].fill      = fl(C_HDR_BG)
    wf["A1"].alignment = ctr()
    wf.row_dimensions[1].height = 22

    for col, label, width in [("A","Topic ID",22),("B","Category",18),
                               ("C","Question",55),("D","Answer",70)]:
        c = wf[f"{col}2"]
        c.value     = label
        c.font      = hf(size=9)
        c.fill      = fl(C_HDR_BG)
        c.alignment = ctr()
        c.border    = bt
        wf.column_dimensions[col].width = width
    wf.row_dimensions[2].height = 20

    frow = 3
    for t in topics:
        cat    = t["category"]
        row_bg = CAT_COLOR.get(cat, "FFFFFF")
        for q, a in t["faqs"]:
            for col, val in zip("ABCD", [t["topic_id"], cat.replace("_"," ").title(), q, a]):
                c = wf[f"{col}{frow}"]
                c.value     = val
                c.font      = cf(size=9)
                c.fill      = fl(row_bg if frow % 2 == 0 else "FFFFFF")
                c.alignment = lft()
                c.border    = bt
            wf.row_dimensions[frow].height = 30
            frow += 1

    # ── Sheet 3: Semester Log ──
    wlog = wb.create_sheet("Semester Log")
    wlog.merge_cells("A1:E1")
    wlog["A1"] = "Semester Update Log"
    wlog["A1"].font      = Font(name="Arial", bold=True, size=11, color=C_HDR_FG)
    wlog["A1"].fill      = fl(C_HDR_BG)
    wlog["A1"].alignment = ctr()
    wlog.row_dimensions[1].height = 22
    for col, label, width in [("A","Semester",16),("B","Updated By",20),
                               ("C","Date",14),("D","Topic Changed",28),("E","Summary",55)]:
        c = wlog[f"{col}2"]
        c.value     = label
        c.font      = hf(size=9)
        c.fill      = fl(C_HDR_BG)
        c.alignment = ctr()
        c.border    = bt
        wlog.column_dimensions[col].width = width
    wlog.row_dimensions[2].height = 20
    for col, val in zip("ABCDE", ["Spring 2026","Admin","2026-01-01",
                                   "All topics","Initial entry from Basic_Info.docx."]):
        c = wlog[f"{col}3"]
        c.value     = val
        c.font      = cf(size=9)
        c.fill      = fl("FFFFFF")
        c.alignment = lft()
        c.border    = bt
    wlog.row_dimensions[3].height = 20

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"  Admin Excel saved → {out_path}")


# ── Main ───────────────────────────────────────────────────────────────────

def build(excel_path: str = None) -> list[dict]:
    """
    Reads basic info data from Google Sheets and returns all chunks.
    excel_path is kept as parameter for compatibility but is ignored.
    """
    import pandas as pd   
    from sheets_reader import read_sheet

    # Read Topics sheet
    df_topics = read_sheet(
        topic      = "basic_info",
        sheet_name = "Basic Info Topics",
        header_row = 2
    )
    df_topics.columns = [c.strip().lower().replace(" ", "_")
                         for c in df_topics.columns]

    # Drop section separator rows
    first_col = df_topics.columns[0]
    df_topics = df_topics[
        pd.to_numeric(df_topics[first_col], errors="coerce").notna()
    ].copy()

    # Normalize column names
    col_map = {}
    for c in df_topics.columns:
        if "topic_id"    in c:               col_map[c] = "topic_id"
        elif "topic_label" in c or "label" in c: col_map[c] = "topic_label"
        elif "category"  in c:               col_map[c] = "category"
        elif "narrative" in c:               col_map[c] = "narrative"
    df_topics = df_topics.rename(columns=col_map)

    # Read FAQ sheet
    df_faqs = read_sheet(
        topic      = "basic_info",
        sheet_name = "FAQ Details",
        header_row = 1
    )
    df_faqs.columns = [c.strip().lower().replace(" ", "_")
                       for c in df_faqs.columns]

    faq_map = {}
    for c in df_faqs.columns:
        if "topic_id"  in c or c == "topic": faq_map[c] = "topic_id"
        elif "question" in c:                faq_map[c] = "question"
        elif "answer"   in c:               faq_map[c] = "answer"
    df_faqs = df_faqs.rename(columns=faq_map)
    df_faqs = df_faqs[df_faqs["topic_id"].notna()]

    # Group FAQs by topic_id
    faq_groups = {}
    for _, row in df_faqs.iterrows():
        tid = str(row["topic_id"]).strip()
        q   = str(row.get("question", "")).strip()
        a   = str(row.get("answer",   "")).strip()
        if q and a:
            faq_groups.setdefault(tid, []).append((q, a))

    # Build chunks
    chunks = []
    for _, row in df_topics.iterrows():
        topic_id    = str(row.get("topic_id",    "")).strip()
        topic_label = str(row.get("topic_label", "")).strip()
        category    = str(row.get("category",    "")).strip().lower().replace(" ", "_")
        narrative   = str(row.get("narrative",   "")).strip()

        if not topic_id or not narrative:
            continue

        base_id = f"{SOURCE}_{SEMESTER}_{topic_id}"
        meta = {
            "topic_id":    topic_id,
            "topic_label": topic_label,
            "category":    category,
            "source":      SOURCE,
            "semester":    SEMESTER,
        }

        chunks.append({
            "chunk_id":   f"{base_id}_narrative",
            "chunk_type": "narrative",
            "topic":      topic_id,
            "semester":   SEMESTER,
            "source":     SOURCE,
            "text":       narrative,
            "metadata":   meta,
        })

        faqs = faq_groups.get(topic_id, [])
        if faqs:
            faq_text = "\n\n".join(f"Q: {q}\nA: {a}" for q, a in faqs)
            chunks.append({
                "chunk_id":   f"{base_id}_faq",
                "chunk_type": "faq",
                "topic":      topic_id,
                "semester":   SEMESTER,
                "source":     SOURCE,
                "text":       faq_text,
                "metadata":   meta,
            })

        chunks.append({
            "chunk_id":   f"{base_id}_metadata",
            "chunk_type": "metadata",
            "topic":      topic_id,
            "semester":   SEMESTER,
            "source":     SOURCE,
            "text": (
                f"Topic: {topic_label} | Category: {category} | "
                f"Source: {SOURCE} | Semester: {SEMESTER}"
            ),
            "metadata": meta,
        })

    return chunks

if __name__ == "__main__":
    build_admin_excel(TOPICS, OUT_XL)

    chunks = build_chunks(TOPICS)

    output = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "semester":     SEMESTER,
        "total_chunks": len(chunks),
        "chunks":       chunks,
    }
    Path(OUT_JSON).parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"  Chunks saved   → {OUT_JSON}")
    print(f"\n  Summary:")
    types = {}
    for c in chunks:
        t = c["chunk_type"]
        types[t] = types.get(t, 0) + 1
    for t, n in sorted(types.items()):
        print(f"    {t:12}: {n}")
    print(f"    {'TOTAL':12}: {len(chunks)}")
