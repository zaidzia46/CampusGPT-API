"""
build_blocks_directory.py
Builds:
  1. blocks_directory_spring_2026.xlsx  — clean merged admin sheet
  2. blocks_chunks.py                   — RAG chunk generator
"""

import json
import re
import argparse
from datetime import datetime, timezone
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Raw data (merged, deduplicated, useful columns only) ──────────────────
# Columns kept: room_id, room_name, room_type, block, floor, wing, adjacent_to, notes
# Dropped: position_order (not useful for RAG), incharge (included in notes where relevant)

ROOMS = [
    # ── A BLOCK — Ground Floor ─────────────────────────────────────────────
    ("A Block", "Ground Floor", "Left wing",  "A3",       "Classroom A3",                  "Classroom",       "Male Toilet",                       "First room on left side of ground floor"),
    ("A Block", "Ground Floor", "Left wing",  "AGF-MT",   "Male Toilet",                   "Toilet",          "Classroom A3",                      "Next to Classroom A3"),
    ("A Block", "Ground Floor", "Left wing",  "AGF-EML",  "Engineering Mechanics Lab",      "Lab",             "Male Toilet",                       "After Male Toilet"),
    ("A Block", "Ground Floor", "Left wing",  "AGF-SR",   "Server Room",                   "Server Room",     "Engineering Mechanics Lab",          "After Engineering Mechanics Lab"),
    ("A Block", "Ground Floor", "Left wing",  "CLAB8",    "Computer Lab 8 (CAD & CAE)",    "Lab",             "Server Room",                       "CAD & CAE Lab, next to Server Room"),
    ("A Block", "Ground Floor", "Left wing",  "AGF-MML",  "Mechanics of Materials Lab",    "Lab",             "Computer Lab 8",                    "Lab Supervisor: Sohail Ahmad"),
    ("A Block", "Ground Floor", "Left wing",  "AGF-TDL",  "Thermodynamics Lab",            "Lab",             "Mechanics of Materials Lab",         "Thermodynamics Lab"),
    ("A Block", "Ground Floor", "Left wing",  "AGF-ICEL", "IC Engine / Engineer Lab",      "Lab",             "Thermodynamics Lab",                 "End of left side"),
    ("A Block", "Ground Floor", "Right wing", "AGF-LIB",  "Library",                       "Library",         "Admin Office",                      "Start of right side"),
    ("A Block", "Ground Floor", "Right wing", "AGF-ADM",  "Admin Office",                  "Office",          "Library, Classroom A1",             "Between Library and Classroom A1"),
    ("A Block", "Ground Floor", "Right wing", "A1",       "Classroom A1",                  "Classroom",       "Admin Office, Female Toilet",        "After Admin Office"),
    ("A Block", "Ground Floor", "Right wing", "AGF-FT",   "Female Toilet",                 "Toilet",          "Classroom A1",                      "Next to Classroom A1"),
    ("A Block", "Ground Floor", "Right wing", "AGF-FFT",  "Female Faculty Toilet",         "Toilet",          "Female Toilet",                     "After Female Toilet"),
    ("A Block", "Ground Floor", "Right wing", "AGF-HR",   "Human Resource Department",     "Department",      "Female Faculty Toilet",              "HR Department"),
    ("A Block", "Ground Floor", "Right wing", "AGF-REG",  "Registration Office",           "Office",          "HR Department, Director Office",    "Near entrance and Director Office"),
    ("A Block", "Ground Floor", "Entrance",   "AGF-DIR",  "Director Office",               "Office",          "Registration Office",               "At the entrance"),
    ("A Block", "Ground Floor", "Entrance",   "AGF-CONF", "Conference Room",               "Conference Room", "Director Office",                   "Near Director Office at entrance"),
    ("A Block", "Ground Floor", "Entrance",   "AGF-FA",   "First Aid Facility Office",     "First Aid",       "Conference Room",                   "First Aid office near entrance"),
    ("A Block", "Ground Floor", "Entrance",   "CLAB9",    "Computer Lab 9",                "Lab",             "First Aid Office",                  "Near entrance side"),
    # ── A BLOCK — First Floor ──────────────────────────────────────────────
    ("A Block", "First Floor",  "Right wing", "A1.1",     "Classroom A1.1",                "Classroom",       "Male Faculty Toilet",               "First room on right side of first floor"),
    ("A Block", "First Floor",  "Right wing", "AFF-MFT",  "Male Faculty Toilet",           "Toilet",          "Classroom A1.1",                    "Next to Classroom A1.1"),
    ("A Block", "First Floor",  "Right wing", "A1.2",     "Classroom A1.2",                "Classroom",       "Male Faculty Toilet",               "After Male Faculty Toilet"),
    ("A Block", "First Floor",  "Right wing", "AFF-RACL", "Refrigeration & Air Conditioning Lab", "Lab",      "Classroom A1.2",                    "R&AC Lab"),
    ("A Block", "First Floor",  "Right wing", "AFF-ACC",  "Accounts Office",               "Office",          "R&AC Lab",                          "Incharge: M. Tauqeer Anwar"),
    ("A Block", "First Floor",  "Right wing", "AFF-PHL",  "Physics Lab",                   "Lab",             "Accounts Office",                   "After Accounts Office"),
    ("A Block", "First Floor",  "Right wing", "AFF-PSD",  "Purchase & Store Department",   "Store",           "Physics Lab",                       "Incharge: M. Nadeem Rana (Senior Program Officer)"),
    ("A Block", "First Floor",  "Right wing", "AFF-SAL",  "Salary Section",                "Office",          "Purchase & Store Department",       "Query hours: 2:30 PM – 4:00 PM"),
    ("A Block", "First Floor",  "Right wing", "CLAB11",   "Computer Lab 11",               "Lab",             "Salary Section",                    "After Salary Section"),
    ("A Block", "First Floor",  "Right wing", "AFF-AUD",  "Auditorium Hall",               "Auditorium",      "Computer Lab 11",                   "Auditorium Hall"),
    ("A Block", "First Floor",  "Right wing", "CLAB12",   "Computer Lab 12",               "Lab",             "Auditorium Hall",                   "After Auditorium"),
    ("A Block", "First Floor",  "Right wing", "AFF-MML",  "Mechanics of Machine Lab",      "Lab",             "Computer Lab 12",                   "Mechanics of Machine Lab"),
    ("A Block", "First Floor",  "Right wing", "AFF-GCR",  "Girls Common Room",             "Common Room",     "Mechanics of Machine Lab",          "Girls Common Room"),
    ("A Block", "First Floor",  "Right wing", "AFF-FML",  "Fluid Mechanics Lab",           "Lab",             "Girls Common Room",                 "Fluid Mechanics Lab"),
    ("A Block", "First Floor",  "Right wing", "AFF-FT",   "Female Toilet",                 "Toilet",          "Fluid Mechanics Lab",               "After Fluid Mechanics Lab"),
    ("A Block", "First Floor",  "Right wing", "A1.3",     "Classroom A1.3",                "Classroom",       "Female Toilet",                     "After Female Toilet"),
    ("A Block", "First Floor",  "Right wing", "AFF-ICL",  "Instrument Control Lab",        "Lab",             "Classroom A1.3",                    "Lab Engineer: Miss Rimsha Sajjad"),
    # ── A BLOCK — Second Floor ─────────────────────────────────────────────
    ("A Block", "Second Floor", "—",          "ASF-FYP",  "FYP Lab (Open Ended Lab)",      "Lab",             "Classroom A2.2",                    "Start of second floor"),
    ("A Block", "Second Floor", "—",          "A2.2",     "Classroom A2.2",                "Classroom",       "FYP Lab",                           "Next to FYP Lab"),
    ("A Block", "Second Floor", "—",          "ASF-MW",   "Male Washroom",                 "Washroom",        "Classroom A2.2",                    "Next to A2.2"),
    ("A Block", "Second Floor", "—",          "ASF-EML",  "Electric Machine Lab",          "Lab",             "Male Washroom",                     "After Male Washroom"),
    ("A Block", "Second Floor", "—",          "ASF-DCO",  "DCO / LDC Office",              "Office",          "Electric Machine Lab",              "DCO / LDC Office"),
    ("A Block", "Second Floor", "—",          "ASF-MFT",  "Male Faculty Toilet",           "Toilet",          "DCO/LDC Office",                    "Next to DCO/LDC Office"),
    ("A Block", "Second Floor", "—",          "A2.1",     "Classroom A2.1",                "Classroom",       "Male Faculty Toilet",               "After Male Faculty Toilet"),
    ("A Block", "Second Floor", "—",          "ASF-CISCO","Cisco Lab",                     "Lab",             "Classroom A2.1",                    "Last room on second floor"),
    # ── B BLOCK — Ground Floor ─────────────────────────────────────────────
    ("B Block", "Ground Floor", "Right wing", "B1",       "Classroom B1",                  "Classroom",       "B2",                                "First room on right side"),
    ("B Block", "Ground Floor", "Right wing", "B2",       "Classroom B2",                  "Classroom",       "B1, B3",                            "Between B1 and B3"),
    ("B Block", "Ground Floor", "Right wing", "B3",       "Classroom B3",                  "Classroom",       "B2, B4",                            "Between B2 and B4"),
    ("B Block", "Ground Floor", "Right wing", "B4",       "Classroom B4",                  "Classroom",       "B3, B5",                            "Between B3 and B5"),
    ("B Block", "Ground Floor", "Right wing", "B5",       "Classroom B5",                  "Classroom",       "B4, B6",                            "Between B4 and B6"),
    ("B Block", "Ground Floor", "Right wing", "B6",       "Classroom B6",                  "Classroom",       "B5, Male Toilet",                   "Last room on right side"),
    ("B Block", "Ground Floor", "Middle",     "BGF-MT",   "Male Toilet",                   "Toilet",          "B6, B7",                            "Between right and middle wings"),
    ("B Block", "Ground Floor", "Middle",     "B7",       "Classroom B7",                  "Classroom",       "Male Toilet, B8",                   "First room in middle section"),
    ("B Block", "Ground Floor", "Middle",     "B8",       "Classroom B8",                  "Classroom",       "B7, B9",                            "Between B7 and B9"),
    ("B Block", "Ground Floor", "Middle",     "B9",       "Classroom B9",                  "Classroom",       "B8, B10",                           "Between B8 and B10"),
    ("B Block", "Ground Floor", "Middle",     "B10",      "Classroom B10",                 "Classroom",       "B9",                                "Last room in middle section"),
    ("B Block", "Ground Floor", "Left wing",  "B11",      "Classroom B11",                 "Classroom",       "B10, B14",                          "First room on left side"),
    ("B Block", "Ground Floor", "Left wing",  "B14",      "Classroom B14",                 "Classroom",       "B11",                               "Last room on left side"),
    # ── C BLOCK — Ground Floor ─────────────────────────────────────────────
    ("C Block", "Ground Floor", "—",          "C1",       "Classroom C1",                  "Classroom",       "—",                                 "First room on ground floor"),
    ("C Block", "Ground Floor", "—",          "C2",       "Classroom C2",                  "Classroom",       "C1",                                "Next to C1"),
    ("C Block", "Ground Floor", "—",          "C3",       "Classroom C3",                  "Classroom",       "C2",                                "Next to C2"),
    ("C Block", "Ground Floor", "Middle",     "GF-HOD",   "HOD Office",                    "Office",          "C3, C4",                            "In the middle of the block, between C3 and C4"),
    ("C Block", "Ground Floor", "Middle",     "GF-DCO",   "DCO Office",                    "Office",          "HOD Office",                        "Near HOD Office"),
    ("C Block", "Ground Floor", "—",          "C4",       "Classroom C4",                  "Classroom",       "HOD Office",                        "After HOD Office"),
    ("C Block", "Ground Floor", "—",          "C5",       "Classroom C5",                  "Classroom",       "Male Faculty Toilet",               "Next to Male Faculty Toilet"),
    ("C Block", "Ground Floor", "—",          "GF-MFT",   "Male Faculty Toilet",           "Toilet",          "C5",                                "End of ground floor, next to C5"),
    # ── C BLOCK — First Floor ──────────────────────────────────────────────
    ("C Block", "First Floor",  "—",          "FF-CR",    "Common Room",                   "Common Room",     "C1.2",                              "Start of first floor"),
    ("C Block", "First Floor",  "—",          "C1.1",     "Classroom C1.1",                "Classroom",       "Common Room",                       "Next to Common Room"),
    ("C Block", "First Floor",  "—",          "C1.2",     "Classroom C1.2",                "Classroom",       "Common Room",                       "Adjacent to Common Room"),
    ("C Block", "First Floor",  "—",          "C1.3",     "Classroom C1.3",                "Classroom",       "C1.2",                              "Next to C1.2"),
    ("C Block", "First Floor",  "—",          "CLAB4",    "Computer Lab 4",                "Lab",             "C1.3",                              "After C1.3"),
    ("C Block", "First Floor",  "—",          "C1.4",     "Classroom C1.4",                "Classroom",       "Computer Lab 4, Computer Lab 3",    "Between CLab4 and CLab3"),
    ("C Block", "First Floor",  "—",          "CLAB3",    "Computer Lab 3",                "Lab",             "C1.4",                              "After C1.4"),
    ("C Block", "First Floor",  "—",          "FF-CONF",  "Conference Room",               "Conference Room", "Computer Lab 3",                    "Next to Computer Lab 3"),
    ("C Block", "First Floor",  "—",          "C1.5",     "Classroom C1.5",                "Classroom",       "Conference Room",                   "After Conference Room"),
    ("C Block", "First Floor",  "—",          "FF-FT",    "Female Toilet",                 "Toilet",          "HEC Smart Classroom",               "Next to HEC Smart Classroom"),
    ("C Block", "First Floor",  "—",          "FF-HEC",   "HEC Smart Classroom",           "Smart Classroom", "Female Toilet",                     "Adjacent to Female Toilet"),
    ("C Block", "First Floor",  "—",          "FF-ITO",   "IT Office",                     "Office",          "HEC Smart Classroom",               "After HEC Smart Classroom"),
    ("C Block", "First Floor",  "—",          "FF-RC",    "Research Center",               "Research Center", "IT Office",                         "Next to IT Office"),
    ("C Block", "First Floor",  "—",          "CLAB2",    "Computer Lab 2",                "Lab",             "Research Center",                   "After Research Center"),
    ("C Block", "First Floor",  "—",          "CLAB1",    "Computer Lab 1",                "Lab",             "Computer Lab 2, Male Toilet",       "Between CLab2 and Male Toilet"),
    ("C Block", "First Floor",  "—",          "FF-MT",    "Male Toilet",                   "Toilet",          "Computer Lab 1, C1.6",              "Next to CLab1"),
    ("C Block", "First Floor",  "—",          "C1.6",     "Classroom C1.6",                "Classroom",       "Male Toilet",                       "Next to Male Toilet"),
    ("C Block", "First Floor",  "—",          "C1.7",     "Classroom C1.7",                "Classroom",       "C1.6",                              "Last classroom on first floor"),
    # ── C BLOCK — Second Floor ─────────────────────────────────────────────
    ("C Block", "Second Floor", "—",          "C2.2",     "Classroom C2.2",                "Classroom",       "—",                                 "First room on second floor"),
    ("C Block", "Second Floor", "—",          "C2.3",     "Classroom C2.3",                "Classroom",       "C2.2",                              "Next to C2.2"),
    ("C Block", "Second Floor", "—",          "CLAB7",    "Computer Lab 7",                "Lab",             "C2.3",                              "After C2.3"),
    ("C Block", "Second Floor", "—",          "C2.4",     "Classroom C2.4",                "Classroom",       "Computer Lab 7",                    "After Computer Lab 7"),
    ("C Block", "Second Floor", "—",          "C2.5",     "Classroom C2.5",                "Classroom",       "Male Toilet",                       "Next to Male Toilet"),
    ("C Block", "Second Floor", "—",          "CLAB6",    "Computer Lab 6",                "Lab",             "C2.5",                              "After C2.5"),
    ("C Block", "Second Floor", "—",          "SF-MT",    "Male Toilet",                   "Toilet",          "C2.5",                              "Next to Classroom C2.5"),
    ("C Block", "Second Floor", "—",          "SF-FYP",   "FYP-SE Office",                 "Office",          "Male Toilet",                       "Last room on second floor"),
]

HEADERS = ["Room ID", "Room Name", "Room Type", "Block", "Floor", "Wing", "Adjacent To", "Notes"]

# ── Excel builder ──────────────────────────────────────────────────────────

def build_excel(out_path: str):
    wb = Workbook()

    C_HDR  = "1F4E79"
    C_FG   = "FFFFFF"
    C_A    = "DEEAF1"
    C_B    = "E2EFDA"
    C_C    = "FFF2CC"
    C_SEC  = "2E75B6"
    C_ALT  = "F5F9FF"
    C_BRD  = "BFBFBF"

    thin = Side(style="thin", color=C_BRD)
    bt   = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hf(bold=True, size=9, color=C_FG):
        return Font(name="Arial", bold=bold, size=size, color=color)
    def cf(size=9):
        return Font(name="Arial", size=size)
    def fl(h):
        return PatternFill("solid", fgColor=h)
    def ctr():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def lft():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    BLOCK_COLOR = {"A Block": C_A, "B Block": C_B, "C Block": C_C}

    ws = wb.active
    ws.title = "All Rooms"

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "COMSATS University Islamabad, Sahiwal Campus — Blocks Directory (A, B, C)"
    ws["A1"].font      = Font(name="Arial", bold=True, size=12, color=C_FG)
    ws["A1"].fill      = fl(C_HDR)
    ws["A1"].alignment = ctr()
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:H2")
    ws["A2"] = "Source of Truth — Spring 2026  |  Edit this file, then run: python generate_chunks.py --semester spring_2026"
    ws["A2"].font      = Font(name="Arial", italic=True, size=8, color="595959")
    ws["A2"].fill      = fl("D9E2F3")
    ws["A2"].alignment = ctr()
    ws.row_dimensions[2].height = 14

    col_widths = [12, 32, 18, 12, 14, 12, 28, 45]
    for i, (h, w) in enumerate(zip(HEADERS, col_widths), 1):
        from openpyxl.utils import get_column_letter
        col = get_column_letter(i)
        c = ws[f"{col}3"]
        c.value     = h
        c.font      = hf()
        c.fill      = fl(C_HDR)
        c.alignment = ctr()
        c.border    = bt
        ws.column_dimensions[col].width = w
    ws.row_dimensions[3].height = 28

    prev_block = None
    prev_floor = None
    excel_row  = 4
    alt        = False

    for (block, floor, wing, room_id, room_name, room_type, adjacent, notes) in ROOMS:
        # Block section header
        if block != prev_block:
            ws.merge_cells(f"A{excel_row}:H{excel_row}")
            c = ws[f"A{excel_row}"]
            c.value     = f"▌ {block}"
            c.font      = Font(name="Arial", bold=True, size=10, color=C_FG)
            c.fill      = fl(C_SEC)
            c.alignment = lft()
            c.border    = bt
            ws.row_dimensions[excel_row].height = 18
            excel_row += 1
            prev_block = block
            prev_floor = None
            alt = False

        # Floor sub-header
        if floor != prev_floor:
            ws.merge_cells(f"A{excel_row}:H{excel_row}")
            c = ws[f"A{excel_row}"]
            c.value     = f"  {floor}"
            c.font      = Font(name="Arial", bold=True, size=9, color="1F4E79")
            c.fill      = fl("EEF4FB")
            c.alignment = lft()
            c.border    = bt
            ws.row_dimensions[excel_row].height = 15
            excel_row += 1
            prev_floor = floor
            alt = False

        row_bg = C_ALT if alt else "FFFFFF"
        alt = not alt

        vals = [room_id, room_name, room_type, block, floor, wing, adjacent, notes]
        for col_idx, val in enumerate(vals, 1):
            from openpyxl.utils import get_column_letter
            c = ws[f"{get_column_letter(col_idx)}{excel_row}"]
            c.value     = val
            c.font      = cf()
            c.fill      = fl(row_bg)
            c.border    = bt
            c.alignment = ctr() if col_idx in (1, 3, 4, 5, 6) else lft()

        ws.row_dimensions[excel_row].height = 22
        excel_row += 1

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:H{excel_row - 1}"

    # ── Summary sheet ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.merge_cells("A1:C1")
    ws2["A1"] = "Block Directory Summary"
    ws2["A1"].font      = Font(name="Arial", bold=True, size=11, color=C_FG)
    ws2["A1"].fill      = fl(C_HDR)
    ws2["A1"].alignment = ctr()
    ws2.row_dimensions[1].height = 22

    for col, label, width in [("A","Block",14),("B","Floor",16),("C","Room Count",14)]:
        c = ws2[f"{col}2"]
        c.value     = label
        c.font      = hf()
        c.fill      = fl(C_HDR)
        c.alignment = ctr()
        c.border    = bt
        ws2.column_dimensions[col].width = width
    ws2.row_dimensions[2].height = 20

    from collections import Counter
    counts = Counter((r[0], r[1]) for r in ROOMS)
    row = 3
    for (block, floor), count in sorted(counts.items()):
        for col, val in zip("ABC", [block, floor, count]):
            c = ws2[f"{col}{row}"]
            c.value     = val
            c.font      = cf()
            c.fill      = fl(BLOCK_COLOR.get(block, "FFFFFF"))
            c.border    = bt
            c.alignment = ctr()
        ws2.row_dimensions[row].height = 18
        row += 1

    # Totals
    ws2[f"B{row}"] = "TOTAL"
    ws2[f"B{row}"].font = hf(color="000000")
    ws2[f"C{row}"] = f"=SUM(C3:C{row-1})"
    ws2[f"C{row}"].font = hf(color="000000")
    for col in "ABC":
        ws2[f"{col}{row}"].border = bt
        ws2[f"{col}{row}"].fill  = fl("D9E2F3")
    ws2.row_dimensions[row].height = 20

    # ── Semester Log ───────────────────────────────────────────────────────
    wl = wb.create_sheet("Semester Log")
    wl.merge_cells("A1:E1")
    wl["A1"] = "Semester Update Log"
    wl["A1"].font      = Font(name="Arial", bold=True, size=11, color=C_FG)
    wl["A1"].fill      = fl(C_HDR)
    wl["A1"].alignment = ctr()
    wl.row_dimensions[1].height = 22
    for col, label, width in [("A","Semester",16),("B","Updated By",20),
                               ("C","Date",14),("D","Block Changed",20),("E","Notes",50)]:
        c = wl[f"{col}2"]
        c.value = label
        c.font  = hf()
        c.fill  = fl(C_HDR)
        c.alignment = ctr()
        c.border = bt
        wl.column_dimensions[col].width = width
    wl.row_dimensions[2].height = 20
    for col, val in zip("ABCDE", ["Spring 2026","Admin","2026-05-08","A, B, C Blocks","Initial entry."]):
        c = wl[f"{col}3"]
        c.value = val
        c.font  = cf()
        c.fill  = fl("FFFFFF")
        c.alignment = lft()
        c.border = bt
    wl.row_dimensions[3].height = 18

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved Excel → {out_path}")

# ── Chunk generator ────────────────────────────────────────────────────────

SOURCE = "blocks_directory"
DEFAULT_SEMESTER = "spring_2026"

def _clean_cell(value) -> str:
    if value is None:
        return ""
    value = str(value).strip()
    return "" if value.lower() in ("none", "nan") else value

def _slug(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", _clean_cell(value).lower()).strip("_")
    return slug or "unknown"

def _normalize_columns(df):
    df = df.copy()
    df.columns = [
        _clean_cell(c).lower().replace("\n", " ").replace("-", " ").replace("_", " ")
        for c in df.columns
    ]

    aliases = {
        "room id": "room_id",
        "room code": "room_id",
        "id": "room_id",
        "room name": "room_name",
        "name": "room_name",
        "room type": "room_type",
        "type": "room_type",
        "block": "block",
        "floor": "floor",
        "wing": "wing",
        "adjacent to": "adjacent_to",
        "adjacent": "adjacent_to",
        "next to": "adjacent_to",
        "notes": "notes",
        "note": "notes",
    }

    return df.rename(columns={c: aliases.get(c, c.replace(" ", "_")) for c in df.columns})

def _rows_from_df(df):
    df = _normalize_columns(df)
    required = ["room_id", "room_name", "room_type", "block", "floor", "wing", "adjacent_to", "notes"]

    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Blocks sheet is missing required columns: {', '.join(missing)}")

    rooms = []
    seen_room_ids = set()
    for _, row in df.iterrows():
        room_id = _clean_cell(row.get("room_id"))
        room_name = _clean_cell(row.get("room_name"))
        block = _clean_cell(row.get("block"))
        floor = _clean_cell(row.get("floor"))

        # Skip generated Excel section rows, blank rows, and other non-record rows.
        if not (room_id and room_name and block and floor):
            continue
        if room_id.startswith("▌") or room_id.lower() in ("room id", "floor", "total"):
            continue

        key = room_id.lower()
        if key in seen_room_ids:
            continue
        seen_room_ids.add(key)

        rooms.append((
            block,
            floor,
            _clean_cell(row.get("wing")) or "-",
            room_id,
            room_name,
            _clean_cell(row.get("room_type")) or "Room",
            _clean_cell(row.get("adjacent_to")) or "-",
            _clean_cell(row.get("notes")),
        ))

    if not rooms:
        raise ValueError("Blocks sheet did not contain any valid room records.")

    return rooms

def build(excel_path: str = None, semester: str = DEFAULT_SEMESTER, **kwargs) -> list[dict]:
    """
    Reads the blocks directory from Google Sheets and returns all chunks.
    excel_path is kept for generator compatibility but is ignored.
    """
    from sheets_reader import read_sheet

    last_error = None
    for header_row in (2, 0, 1):
        try:
            df = read_sheet(
                topic      = "blocks_directory",
                sheet_name = "All Rooms",
                header_row = header_row
            )
            rooms = _rows_from_df(df)
            break
        except ValueError as exc:
            last_error = exc
    else:
        raise last_error

    print(f"[DEBUG] Loaded {len(rooms)} block directory records from Google Sheets")
    return build_chunks(rooms, semester=semester)

def build_chunks(rooms=None, semester: str = DEFAULT_SEMESTER, **kwargs) -> list[dict]:
    if rooms is None:
        rooms = ROOMS  # fallback to hardcoded for direct script runs

    chunks = []

    # ── 1. Per-room narrative + FAQ chunks
    for (block, floor, wing, room_id, room_name, room_type, adjacent, notes) in rooms:
        slug    = _slug(room_id)
        base_id = f"{SOURCE}_{semester}_{_slug(block)}_{slug}"

        wing_text = f" ({wing})" if wing and wing not in ("—", "-") else ""
        adj_text  = f" It is adjacent to {adjacent}." if adjacent not in ("—", "-", "") else ""

        narrative = (
            f"{room_name} (Room ID: {room_id}) is located in {block}, {floor}{wing_text} "
            f"at COMSATS University Islamabad, Sahiwal Campus.{adj_text} "
            f"Room type: {room_type}. {notes}."
        )

        faq = "\n\n".join([
            f"Q: Where is {room_name} in COMSATS Sahiwal?\nA: {room_name} is on the {floor} of {block}{wing_text}. {notes}.",
            f"Q: What is in room {room_id} at COMSATS Sahiwal?\nA: Room {room_id} is {room_name}, a {room_type} on the {floor} of {block}.",
            f"Q: What is next to {room_name}?\nA: {room_name} is adjacent to {adjacent}." if adjacent not in ("—","-","") else
            f"Q: What floor is {room_name} on?\nA: {room_name} is on the {floor} of {block}.",
        ])

        meta = {
            "room_id":   room_id,
            "room_name": room_name,
            "room_type": room_type,
            "block":     block,
            "floor":     floor,
            "wing":      wing,
            "source":    SOURCE,
            "semester":  semester,
        }

        chunks.append({
            "chunk_id":   f"{base_id}_narrative",
            "chunk_type": "narrative",
            "topic":      SOURCE,
            "semester":   semester,
            "source":     SOURCE,
            "text":       narrative,
            "metadata":   meta,
        })
        chunks.append({
            "chunk_id":   f"{base_id}_faq",
            "chunk_type": "faq",
            "topic":      SOURCE,
            "semester":   semester,
            "source":     SOURCE,
            "text":       faq,
            "metadata":   meta,
        })

    # ── 2. Floor-level summary chunks (one per block+floor combo) ─────────
    from collections import defaultdict
    floor_groups = defaultdict(list)
    for row in rooms:
        key = (row[0], row[1])  # (block, floor)
        floor_groups[key].append(row)

    for (block, floor), rooms in floor_groups.items():
        slug    = f"{_slug(block)}_{_slug(floor)}"
        base_id = f"{SOURCE}_{semester}_{slug}_overview"

        # Group by type
        by_type = defaultdict(list)
        for r in rooms:
            by_type[r[5]].append(r[4])  # room_type → [room_names]

        type_lines = []
        for rtype, names in sorted(by_type.items()):
            type_lines.append(f"{rtype}s: {', '.join(names)}")

        narrative = (
            f"The {floor} of {block} at COMSATS University Islamabad, Sahiwal Campus "
            f"contains {len(rooms)} rooms. "
            + " | ".join(type_lines) + "."
        )

        faq = "\n\n".join([
            f"Q: What rooms are on the {floor} of {block}?\n"
            f"A: The {floor} of {block} has {len(rooms)} rooms including: "
            + ", ".join(r[4] for r in rooms) + ".",

            f"Q: What labs are on the {floor} of {block}?\n"
            f"A: " + (
                "Labs on this floor: " + ", ".join(r[4] for r in rooms if r[5] == "Lab") + "."
                if any(r[5] == "Lab" for r in rooms)
                else "There are no labs on this floor."
            ),

            f"Q: What offices are in {block} on the {floor}?\n"
            f"A: " + (
                "Offices on this floor: " + ", ".join(r[4] for r in rooms if r[5] == "Office") + "."
                if any(r[5] == "Office" for r in rooms)
                else "There are no offices on this floor."
            ),
        ])

        chunks.append({
            "chunk_id":   f"{base_id}_narrative",
            "chunk_type": "narrative",
            "topic":      SOURCE,
            "semester":   semester,
            "source":     SOURCE,
            "text":       narrative,
            "metadata":   {"block": block, "floor": floor, "source": SOURCE, "semester": semester},
        })
        chunks.append({
            "chunk_id":   f"{base_id}_faq",
            "chunk_type": "faq",
            "topic":      SOURCE,
            "semester":   semester,
            "source":     SOURCE,
            "text":       faq,
            "metadata":   {"block": block, "floor": floor, "source": SOURCE, "semester": semester},
        })

    # ── 3. Block-level overview chunks ────────────────────────────────────
    block_groups = defaultdict(list)
    for row in rooms:
        block_groups[row[0]].append(row)

    BLOCK_DEPT = {
        "A Block": "Engineering Department",
        "B Block": "Management Sciences Department",
        "C Block": "Computer Science Department",
    }

    for block, rooms in block_groups.items():
        slug    = _slug(block)
        base_id = f"{SOURCE}_{semester}_{slug}_overview"
        dept    = BLOCK_DEPT.get(block, "")
        floors  = ", ".join(dict.fromkeys(r[1] for r in rooms))

        notable = [r[4] for r in rooms if r[5] not in ("Classroom","Toilet","Washroom")]
        notable_str = ", ".join(notable[:10]) if notable else "classrooms on multiple floors"

        narrative = (
            f"{block} at COMSATS University Islamabad, Sahiwal Campus"
            f"{f' houses the {dept}' if dept else ''}. "
            f"It has {len(rooms)} rooms across {floors}. "
            f"Notable facilities include: {notable_str}."
        )

        faq = "\n\n".join([
            f"Q: What is in {block} at COMSATS Sahiwal?\n"
            f"A: {block}{f' houses the {dept}' if dept else ''}. It has {len(rooms)} rooms including classrooms, labs, offices and facilities.",

            f"Q: Which block has the {dept}?\n"
            f"A: The {dept} is in {block}." if dept else
            f"Q: How many rooms are in {block}?\nA: {block} has {len(rooms)} rooms across {floors}.",

            f"Q: What labs are in {block}?\n"
            f"A: Labs in {block}: " + ", ".join(r[4] for r in rooms if r[5] == "Lab") + ".",

            f"Q: What offices are in {block}?\n"
            f"A: Offices in {block}: " + ", ".join(r[4] for r in rooms if r[5] == "Office") + ".",
        ])

        chunks.append({
            "chunk_id":   f"{base_id}_narrative",
            "chunk_type": "narrative",
            "topic":      SOURCE,
            "semester":   semester,
            "source":     SOURCE,
            "text":       narrative,
            "metadata":   {"block": block, "source": SOURCE, "semester": semester},
        })
        chunks.append({
            "chunk_id":   f"{base_id}_faq",
            "chunk_type": "faq",
            "topic":      SOURCE,
            "semester":   semester,
            "source":     SOURCE,
            "text":       faq,
            "metadata":   {"block": block, "source": SOURCE, "semester": semester},
        })

    return chunks


# ── Main ───────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--semester", default=DEFAULT_SEMESTER)
    args = parser.parse_args()

    semester = args.semester
    xl_path     = f"/home/claude/cui_rag/data/sources/blocks_directory_{semester}.xlsx"
    chunks_path = f"/home/claude/cui_rag/data/chunks/blocks_chunks_{semester}.json"

    build_excel(xl_path)

    chunks = build_chunks(semester=semester)
    output = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "semester":     semester,
        "total_chunks": len(chunks),
        "chunks":       chunks,
    }
    Path(chunks_path).parent.mkdir(parents=True, exist_ok=True)
    with open(chunks_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"Saved {len(chunks)} chunks → {chunks_path}")
    by_type = {}
    for c in chunks:
        t = c["chunk_type"]
        by_type[t] = by_type.get(t, 0) + 1
    for t, n in sorted(by_type.items()):
        print(f"  {t:12}: {n}")
